"""
Ynov'iT Presales Pipeline — Générateur Chiffrage Excel

Transforme la sortie de l'agent chiffrage en fichier Excel
avec un onglet synthèse et un onglet détail.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("presales.generators.xlsx")

# Couleurs
NAVY_FILL = PatternFill(start_color="1A2B4A", end_color="1A2B4A", fill_type="solid")
TEAL_FILL = PatternFill(start_color="3EC9A7", end_color="3EC9A7", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F0F2F5", end_color="F0F2F5", fill_type="solid")
LIGHT_TEAL = PatternFill(start_color="D1F7EE", end_color="D1F7EE", fill_type="solid")
LIGHT_ORANGE = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1A2B4A")
BODY_FONT = Font(name="Arial", size=10, color="333333")
BOLD_FONT = Font(name="Arial", size=10, bold=True, color="1A2B4A")
TOTAL_FONT = Font(name="Arial", size=12, bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style="thin", color="E2E6EA"),
    right=Side(style="thin", color="E2E6EA"),
    top=Side(style="thin", color="E2E6EA"),
    bottom=Side(style="thin", color="E2E6EA"),
)


def generate_chiffrage_xlsx(chiffrage_data: dict, output_path: Path, societe: dict = None):
    """
    Génère le chiffrage en fichier Excel.

    Args:
        chiffrage_data: Sortie de l'agent chiffrage (uo_brut + ajustement)
        output_path: Chemin du fichier .xlsx à créer
        societe: Infos société pour le titre
    """
    wb = Workbook()

    # ── Onglet Synthèse ────────────────────────────────────
    ws_synth = wb.active
    ws_synth.title = "Synthèse"
    _build_synthese(ws_synth, chiffrage_data, societe)

    # ── Onglet Détail ──────────────────────────────────────
    ws_detail = wb.create_sheet("Détail par ligne")
    _build_detail(ws_detail, chiffrage_data)

    # ── Onglet Risques ─────────────────────────────────────
    ws_risques = wb.create_sheet("Risques")
    _build_risques(ws_risques, chiffrage_data)

    # ── Sauvegarde ─────────────────────────────────────────
    wb.save(str(output_path))
    logger.info(f"Chiffrage Excel généré : {output_path}")


def _build_synthese(ws, data, societe):
    """Construit l'onglet synthèse par module."""
    nom = societe.get("raison_sociale", "Prospect") if societe else "Prospect"

    # Titre
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Chiffrage projet Odoo 19 — {nom}"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 35

    # En-têtes
    headers = ["Module", "Statut", "UO Brut (j)", "Coefficient", "UO Ajusté (j)", "UO Final (j)", "Justification"]
    widths = [30, 10, 14, 12, 14, 14, 55]

    row = 3
    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[row].height = 30

    # Données par module
    ajustement = data.get("ajustement", {})
    par_module = ajustement.get("par_module", {})
    uo_brut_data = data.get("uo_brut", {}).get("par_module", {})

    row = 4
    for mod_id, mod_info in sorted(par_module.items()):
        label = mod_info.get("label", mod_id)
        uo_brut = mod_info.get("uo_brut", 0)
        coeff = mod_info.get("coefficient", 1.0)
        uo_ajuste = mod_info.get("uo_ajuste", 0)
        justif = mod_info.get("justification", "")

        # Couleur de fond alternée
        fill = LIGHT_GRAY if row % 2 == 0 else None

        ws.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row, column=2, value="Actif").font = BODY_FONT
        ws.cell(row=row, column=3, value=uo_brut).font = BODY_FONT
        ws.cell(row=row, column=3).number_format = "0.0"
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=4, value=coeff).font = BODY_FONT
        ws.cell(row=row, column=4).number_format = "0.00"
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")
        uo_final = mod_info.get("uo_final", uo_ajuste)

        ws.cell(row=row, column=5, value=uo_ajuste).font = BODY_FONT
        ws.cell(row=row, column=5).number_format = "0.0"
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=6, value=uo_final).font = BOLD_FONT
        ws.cell(row=row, column=6).number_format = "0.0"
        ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=7, value=justif).font = BODY_FONT
        ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True)

        # Colorer le coefficient si != 1.0
        if coeff > 1.0:
            ws.cell(row=row, column=4).fill = LIGHT_ORANGE

        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
            if fill:
                ws.cell(row=row, column=col).fill = fill

        ws.row_dimensions[row].height = 40
        row += 1

    # Ligne de total
    row += 1
    ws.merge_cells(f"A{row}:D{row}")
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TEAL_FILL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")

    total_ajuste = ajustement.get("total_uo_ajuste", 0)
    ws.cell(row=row, column=5, value=total_ajuste).font = TOTAL_FONT
    ws.cell(row=row, column=5).fill = TEAL_FILL
    ws.cell(row=row, column=5).number_format = "0.0"
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="center")

    total_final = ajustement.get("total_uo_final", total_ajuste)
    ws.cell(row=row, column=6, value=total_final).font = TOTAL_FONT
    ws.cell(row=row, column=6).fill = TEAL_FILL
    ws.cell(row=row, column=6).number_format = "0.0"
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center")

    ecart = ajustement.get("ecart_global_pct", 0)
    ws.cell(row=row, column=7, value=f"Écart : +{ecart}% vs brut").font = TOTAL_FONT
    ws.cell(row=row, column=7).fill = TEAL_FILL

    for col in range(1, 8):
        ws.cell(row=row, column=col).border = THIN_BORDER

    ws.row_dimensions[row].height = 30

    # Bloc catégorie prospect
    categorie = data.get("categorie", {})
    if categorie:
        row += 2
        ws.merge_cells(f"A{row}:G{row}")
        ws.cell(row=row, column=1, value="Coefficient de catégorie prospect").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_GRAY
        row += 1

        coeff_cat = categorie.get("coefficient_combine", 1.0)
        taille = categorie.get("taille", {})
        sante = categorie.get("sante", {})
        resume = categorie.get("resume", "")

        cat_rows = [
            ("Catégorie taille", f"{taille.get('label', '')}  {taille.get('detail', '')}  (×{taille.get('coefficient', 1.0)})"),
            ("Santé financière", f"{sante.get('label', '')}  {sante.get('detail', '')}  (×{sante.get('coefficient', 1.0)})"),
            ("Coefficient combiné", f"×{coeff_cat}  —  {resume}"),
        ]
        for label, val in cat_rows:
            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.merge_cells(f"B{row}:G{row}")
            ws.cell(row=row, column=2, value=val).font = BODY_FONT
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = THIN_BORDER
            row += 1

    # Justification globale
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value="Justification globale :").font = BOLD_FONT
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    ws.cell(row=row, column=1, value=ajustement.get("justification_globale", "")).font = BODY_FONT
    ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
    ws.row_dimensions[row].height = 60


def _build_detail(ws, data):
    """Construit l'onglet détail ligne par ligne."""
    headers = ["Question", "Module", "Type UO", "Valeur (j)", "Détail"]
    widths = [45, 18, 14, 12, 40]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    lignes = data.get("uo_brut", {}).get("lignes", [])
    for i, line in enumerate(lignes, 2):
        ws.cell(row=i, column=1, value=line.get("label", "")).font = BODY_FONT
        mod = (line.get("module") or "Général").replace("has_", "")
        ws.cell(row=i, column=2, value=mod).font = BODY_FONT
        ws.cell(row=i, column=3, value=line.get("uo_type", "")).font = BODY_FONT
        ws.cell(row=i, column=4, value=line.get("uo_value", 0)).font = BODY_FONT
        ws.cell(row=i, column=4).number_format = "0.0"
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=5, value=line.get("detail", "")).font = BODY_FONT

        fill = LIGHT_GRAY if i % 2 == 0 else None
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER
            if fill:
                ws.cell(row=i, column=col).fill = fill

    # Total en bas
    total_row = len(lignes) + 2
    ws.cell(row=total_row, column=3, value="TOTAL").font = BOLD_FONT
    ws.cell(row=total_row, column=4).font = BOLD_FONT
    ws.cell(row=total_row, column=4).number_format = "0.0"
    ws.cell(row=total_row, column=4, value=data.get("uo_brut", {}).get("total_uo", 0))
    ws.cell(row=total_row, column=4).fill = LIGHT_TEAL


def _build_risques(ws, data):
    """Construit l'onglet risques."""
    headers = ["Module", "Description", "Impact (j)", "Probabilité"]
    widths = [18, 60, 12, 14]

    for col, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = NAVY_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    risques = data.get("ajustement", {}).get("risques", [])
    for i, risque in enumerate(risques, 2):
        ws.cell(row=i, column=1, value=risque.get("module", "global")).font = BODY_FONT
        ws.cell(row=i, column=2, value=risque.get("description", "")).font = BODY_FONT
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=3, value=risque.get("impact_uo", 0)).font = BODY_FONT
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=4, value=risque.get("probabilite", "")).font = BODY_FONT
        ws.cell(row=i, column=4).alignment = Alignment(horizontal="center")

        for col in range(1, 5):
            ws.cell(row=i, column=col).border = THIN_BORDER

        ws.row_dimensions[i].height = 35
